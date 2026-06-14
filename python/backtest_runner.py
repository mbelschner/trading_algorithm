"""
Walk-Forward Backtester mit Plugin-Strategien, Parallel Processing
und Excel-Output.

KONFIGURATION: siehe CONFIG-Block weiter unten. Einfach das Script
aus dem Editor starten (F5). CLI-Args ueberschreiben CONFIG nur,
wenn explizit gesetzt.
"""
from __future__ import annotations

import argparse
import glob
import importlib.util
import os
import sys
import time
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, asdict
from pathlib import Path

import numpy as np
import pandas as pd

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Konfiguration
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
STRATEGIES_DIR = SCRIPT_DIR / "strategies"
PRICE_DATA_DIR = SCRIPT_DIR.parent / "price_data"
OUTPUT_DIR = SCRIPT_DIR / "results"
OUTPUT_DIR.mkdir(exist_ok=True)

IS_SPLIT = 0.70  # 70% In-Sample, 30% Out-of-Sample


# ===========================================================================
# >>> HIER EINSTELLEN <<<  (None = "alle nehmen")
# ===========================================================================
CONFIG = {
    # Liste von CSV-Dateinamen (mit .csv) aus dem price_data/-Ordner.
    # None = alle .csv im Ordner verwenden.
    # Beispiel: ["GOLD_MINUTE_5.csv", "US500_MINUTE_5.csv", "DE40_MINUTE_15.csv"]
    "instruments": "GOLD_MINUTE_5.csv",

    # Liste von Strategie-Dateinamen (ohne .py) aus dem strategies/-Ordner.
    # None = alle Strategien laden.
    # Beispiel: ["rsi_mean_reversion", "donchian_breakout"]
    "strategies": None,

    # Anzahl Worker-Prozesse. None = cpu_count - 1.
    "workers": None,

    # Transaktionskosten pro Side in Basispunkten (Round-Trip = 2x).
    # Capital.com Spreads grob: Gold ~3-5bp, US500 ~2-3bp, Brent ~5-8bp.
    "cost_bps": 1.0,

    # Output-Pfad fuer Excel. None = automatisch in results/ mit Timestamp.
    "output": None,
}
# ===========================================================================


# ---------------------------------------------------------------------------
# Strategy Loader
# ---------------------------------------------------------------------------
def load_strategies(only: list[str] | None = None) -> dict:
    """Laedt alle Strategie-Module aus strategies/."""
    strategies = {}
    for path in sorted(STRATEGIES_DIR.glob("*.py")):
        if path.stem.startswith("_"):
            continue
        if only and path.stem not in only and (not any(o in path.stem for o in only)):
            continue
        spec = importlib.util.spec_from_file_location(path.stem, path)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        if not hasattr(mod, "generate_signals") or not hasattr(mod, "PARAM_GRID"):
            print(f"[WARN] {path.name} hat kein gueltiges Strategie-Interface - uebersprungen.")
            continue
        strategies[getattr(mod, "NAME", path.stem)] = {
            "module_path": str(path),
            "param_grid": mod.PARAM_GRID,
        }
    return strategies


def _import_strategy_from_path(module_path: str):
    """Wird im Worker-Prozess gerufen (kein Pickling von Modulen noetig)."""
    path = Path(module_path)
    spec = importlib.util.spec_from_file_location(path.stem, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def expand_grid(grid: dict):
    from itertools import product
    if not grid:
        yield {}
        return
    keys = list(grid.keys())
    for values in product(*[grid[k] for k in keys]):
        yield dict(zip(keys, values))


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------
def load_price_data(filepath: str) -> pd.DataFrame:
    df = pd.read_csv(filepath, parse_dates=["time"])
    df = df.rename(columns={"time": "Timestamp", "close": "Close", "open": "Open",
                            "high": "High", "low": "Low", "volume": "Volume"})
    df = df.set_index("Timestamp").sort_index()
    df = df[~df.index.duplicated(keep="last")]
    return df


def infer_bars_per_year(df: pd.DataFrame) -> int:
    """Schaetzt bars_per_year aus den tatsaechlichen Zeitdifferenzen."""
    if len(df) < 2:
        return 252
    deltas = df.index.to_series().diff().dropna()
    median_delta = deltas.median()
    minutes = max(median_delta.total_seconds() / 60.0, 0.1)
    if minutes >= 60 * 20:  # Daily oder groesser
        return 252
    bars_per_day = (60 * 24) / minutes  # 24h Markt (Forex/Commodities/Index-CFDs)
    return int(bars_per_day * 252)


# ---------------------------------------------------------------------------
# Backtest Core
# ---------------------------------------------------------------------------
@dataclass
class Metrics:
    sharpe: float = 0.0
    max_dd: float = 0.0
    num_trades: int = 0
    win_rate: float = 0.0
    profit_factor: float = 0.0
    total_return: float = 0.0
    exposure: float = 0.0  # Anteil bars in market


def compute_metrics(df: pd.DataFrame, bars_per_year: int, cost_bps: float = 0.0) -> Metrics:
    """
    cost_bps: Kosten pro Side in Basispunkten. Bei einem Round-Trip werden
    also 2*cost_bps angezogen. Wird auf Bars angewendet, an denen sich
    die Position aendert.
    """
    if "Position" not in df.columns or len(df) == 0:
        return Metrics()

    pos = df["Position"].fillna(0).astype(float)
    market_ret = np.log(df["Close"] / df["Close"].shift(1))

    # Position der vorigen Bar bestimmt Return der aktuellen (kein Look-ahead)
    pos_lag = pos.shift(1).fillna(0)
    strat_ret = pos_lag * market_ret

    # Transaktionskosten: jede Aenderung der Position kostet
    pos_change = pos.diff().abs().fillna(0)
    cost = pos_change * (cost_bps / 10_000.0)
    strat_ret = strat_ret - cost

    strat_ret = strat_ret.dropna()
    if len(strat_ret) == 0:
        return Metrics()

    mean_r = strat_ret.mean()
    std_r = strat_ret.std()
    sharpe = float((mean_r / std_r) * np.sqrt(bars_per_year)) if std_r > 0 else 0.0

    cum = np.exp(strat_ret.cumsum())
    max_dd = float(((cum - cum.cummax()) / cum.cummax()).min())
    total_return = float(np.exp(strat_ret.sum()) - 1)

    # Trade-Ebene
    pos_lag_aligned = pos_lag.reindex(strat_ret.index).fillna(0)
    trade_id = (pos_lag_aligned != pos_lag_aligned.shift(1)).cumsum()
    trade_rets = strat_ret.groupby(trade_id).sum()
    trade_pos = pos_lag_aligned.groupby(trade_id).first()
    active = trade_rets[trade_pos != 0]

    num_trades = int(len(active))
    if num_trades > 0:
        win_rate = float((active > 0).sum() / num_trades)
        gp = float(active[active > 0].sum())
        gl = float(abs(active[active < 0].sum()))
        profit_factor = (gp / gl) if gl > 0 else float("inf")
    else:
        win_rate = 0.0
        profit_factor = 0.0

    exposure = float((pos_lag_aligned != 0).mean())

    return Metrics(sharpe, max_dd, num_trades, win_rate, profit_factor, total_return, exposure)


def apply_filters(is_m: Metrics, os_m: Metrics) -> dict:
    f = {
        "F1_OS_DD_above_-35pct":      os_m.max_dd > -0.35,
        "F2_OS_Sharpe_gt_0.5":        os_m.sharpe > 0.5,
        "F3_OS_Sharpe_lt_2.5":        os_m.sharpe < 2.5,
        "F4_OS_le_IS_x1.3":           os_m.sharpe <= is_m.sharpe * 1.30,
        "F5_OS_Trades_ge_30":         os_m.num_trades >= 30,
        "F6_IS_Sharpe_positive":      is_m.sharpe > 0,
    }
    f["PASSED_ALL"] = all(f.values())
    return f


# ---------------------------------------------------------------------------
# Worker (laeuft im Subprozess)
# ---------------------------------------------------------------------------
def run_single_backtest(task: dict) -> dict:
    """Ein Task = (Strategie, Instrument, Parameter)."""
    try:
        mod = _import_strategy_from_path(task["module_path"])
        df = load_price_data(task["data_path"])
        bars_per_year = infer_bars_per_year(df)

        split_idx = int(len(df) * IS_SPLIT)
        is_df = df.iloc[:split_idx]
        os_df = df.iloc[split_idx:]

        is_sig = mod.generate_signals(is_df, **task["params"])
        os_sig = mod.generate_signals(os_df, **task["params"])

        is_m = compute_metrics(is_sig, bars_per_year, cost_bps=task["cost_bps"])
        os_m = compute_metrics(os_sig, bars_per_year, cost_bps=task["cost_bps"])
        filters = apply_filters(is_m, os_m)

        row = {
            "Strategy":   task["strategy_name"],
            "Instrument": task["instrument"],
            "Params":     str(task["params"]),
            **{f"Param_{k}": v for k, v in task["params"].items()},
            "Bars_per_Year": bars_per_year,
            "Bars_IS": len(is_df),
            "Bars_OS": len(os_df),
            # IS
            "IS_Sharpe":       round(is_m.sharpe, 3),
            "IS_MaxDD":        round(is_m.max_dd, 4),
            "IS_TotalReturn":  round(is_m.total_return, 4),
            "IS_WinRate":      round(is_m.win_rate, 4),
            "IS_ProfitFactor": round(is_m.profit_factor, 3) if np.isfinite(is_m.profit_factor) else None,
            "IS_Trades":       is_m.num_trades,
            "IS_Exposure":     round(is_m.exposure, 3),
            # OS
            "OS_Sharpe":       round(os_m.sharpe, 3),
            "OS_MaxDD":        round(os_m.max_dd, 4),
            "OS_TotalReturn":  round(os_m.total_return, 4),
            "OS_WinRate":      round(os_m.win_rate, 4),
            "OS_ProfitFactor": round(os_m.profit_factor, 3) if np.isfinite(os_m.profit_factor) else None,
            "OS_Trades":       os_m.num_trades,
            "OS_Exposure":     round(os_m.exposure, 3),
            # Filter
            **filters,
            "Error": None,
        }
        return row
    except Exception as e:
        return {
            "Strategy": task.get("strategy_name"),
            "Instrument": task.get("instrument"),
            "Params": str(task.get("params")),
            "Error": f"{type(e).__name__}: {e}",
            "PASSED_ALL": False,
        }


# ---------------------------------------------------------------------------
# Excel Output
# ---------------------------------------------------------------------------
def write_excel(results_df: pd.DataFrame, output_path: Path):
    """Schreibt Excel mit drei Sheets: All, Passed, Summary."""
    passed = results_df[results_df.get("PASSED_ALL", False) == True].copy()
    passed = passed.sort_values("OS_Sharpe", ascending=False) if "OS_Sharpe" in passed.columns else passed

    agg_kwargs: dict = dict(Runs=("Strategy", "size"), Passed=("PASSED_ALL", "sum"))
    if "OS_Sharpe" in results_df.columns:
        agg_kwargs["Best_OS_Sharpe"] = ("OS_Sharpe", "max")
        agg_kwargs["Median_OS_Sharpe"] = ("OS_Sharpe", "median")
    if "IS_Sharpe" in results_df.columns:
        agg_kwargs["Median_IS_Sharpe"] = ("IS_Sharpe", "median")
    summary = (results_df
               .groupby(["Strategy", "Instrument"])
               .agg(**agg_kwargs)
               .reset_index())
    summary["Pass_Rate"] = (summary["Passed"] / summary["Runs"]).round(3)

    with pd.ExcelWriter(output_path, engine="openpyxl") as w:
        results_df.to_excel(w, sheet_name="All_Runs", index=False)
        passed.to_excel(w, sheet_name="Passed_Filter", index=False)
        summary.to_excel(w, sheet_name="Summary", index=False)

        # leichte Formatierung
        from openpyxl.styles import PatternFill, Font
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        bold = Font(bold=True)

        for sheet in ("All_Runs", "Passed_Filter", "Summary"):
            ws = w.sheets[sheet]
            for cell in ws[1]:
                cell.font = bold
            if sheet == "All_Runs" and "PASSED_ALL" in results_df.columns:
                col_idx = list(results_df.columns).index("PASSED_ALL") + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = green if cell.value is True else red


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser()
    p.add_argument("--instruments", nargs="*", default=None,
                   help="CSV-Dateien aus price_data/. Ueberschreibt CONFIG.")
    p.add_argument("--strategies", nargs="*", default=None,
                   help="Strategie-Dateinamen ohne .py. Ueberschreibt CONFIG.")
    p.add_argument("--workers", type=int, default=None)
    p.add_argument("--cost-bps", type=float, default=None,
                   help="Transaktionskosten pro Side in bp.")
    p.add_argument("--output", type=str, default=None)
    args = p.parse_args()

    # CLI > CONFIG > Default
    instruments = args.instruments if args.instruments is not None else CONFIG["instruments"]
    # Normalize: CONFIG may hold a bare string instead of a list
    if isinstance(instruments, str):
        instruments = [instruments]
    strategies_filter = args.strategies if args.strategies is not None else CONFIG["strategies"]
    workers = args.workers if args.workers is not None else (CONFIG["workers"] or max(1, os.cpu_count() - 1))
    cost_bps = args.cost_bps if args.cost_bps is not None else CONFIG["cost_bps"]
    output = args.output if args.output is not None else CONFIG["output"]

    # Strategien
    strategies = load_strategies(strategies_filter)
    if not strategies:
        print("Keine Strategien gefunden."); return
    print(f"Geladene Strategien: {list(strategies.keys())}")

    # Instrumente
    if instruments:
        files = [PRICE_DATA_DIR / f for f in instruments]
    else:
        files = sorted(PRICE_DATA_DIR.glob("*.csv"))
    files = [f for f in files if Path(f).exists()]
    if not files:
        print(f"Keine CSV-Dateien gefunden in {PRICE_DATA_DIR}"); return
    print(f"Instrumente: {[f.name for f in files]}")

    # Task-Liste aufbauen
    tasks = []
    for strat_name, meta in strategies.items():
        for params in expand_grid(meta["param_grid"]):
            for f in files:
                tasks.append({
                    "strategy_name": strat_name,
                    "module_path":   meta["module_path"],
                    "instrument":    Path(f).stem,
                    "data_path":     str(f),
                    "params":        params,
                    "cost_bps":      cost_bps,
                })

    print(f"Gesamt Tasks: {len(tasks)}  |  Worker: {workers}  |  Kosten: {cost_bps} bp/side")
    t0 = time.time()

    results = []
    with ProcessPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(run_single_backtest, t) for t in tasks]
        for i, fut in enumerate(as_completed(futures), 1):
            results.append(fut.result())
            if i % 25 == 0 or i == len(tasks):
                print(f"  fertig: {i}/{len(tasks)}  ({(time.time()-t0):.1f}s)")

    results_df = pd.DataFrame(results)

    # Output
    out = Path(output) if output else OUTPUT_DIR / f"backtest_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    write_excel(results_df, out)
    print(f"\n[OK] Excel geschrieben: {out}")
    print(f"Passed Filter: {int(results_df.get('PASSED_ALL', pd.Series(dtype=bool)).sum())}/{len(results_df)}")


if __name__ == "__main__":
    main()